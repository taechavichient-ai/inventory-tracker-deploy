import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import os
_HERE = os.path.dirname(os.path.abspath(__file__))
BASE  = os.environ.get('DATA_DIR',
        os.path.join(_HERE, 'mnt', 'ตรวจสอบสินค้าจัดส่งสาขา')) + os.sep
DB    = os.path.join(BASE.rstrip(os.sep), 'database') + os.sep
def load_db(db_path, src_path, header_row=1):
    """Load from database CSV if exists, else fall back to source xlsx."""
    if os.path.exists(db_path):
        df = pd.read_csv(db_path, dtype=str)
        print(f'  DB: {db_path} ({len(df):,} rows)')
    else:
        raw = pd.read_excel(src_path, header=header_row)
        df = raw[raw.iloc[:, 0] != 'Total'].copy()
        print(f'  SRC fallback: {src_path} ({len(df):,} rows)')
    return df

tr = load_db(DB + 'tr_database.csv', BASE + 'รายละเอียดการร้องขอโอนวัตถุดิบ.xlsx')
to = load_db(DB + 'to_database.csv', BASE + 'รายละเอียดการโอนวัตถุดิบ.xlsx')
gr = load_db(DB + 'gr_database.csv', BASE + 'รายละเอียดการรับโอนวัตถุดิบ.xlsx')

# Ensure numeric qty columns are numeric
for col in ['ขอโอนสินค้า']:
    if col in tr.columns: tr[col] = pd.to_numeric(tr[col], errors='coerce')
for col in ['โอนสินค้า']:
    if col in to.columns: to[col] = pd.to_numeric(to[col], errors='coerce')
for col in ['เติมสินค้า']:
    if col in gr.columns: gr[col] = pd.to_numeric(gr[col], errors='coerce')

# --- Aggregate TO by TO + product ---
to_agg = to.groupby(['TO', 'รหัสวัตถุดิบ'], as_index=False).agg(
    วันที่_TO=('วันที่ออก TO', 'first'),
    TR_from_TO=('TR', 'first'),
    GR_from_TO=('GR', 'first'),
    ชื่อ_TO=('ชื่อ', 'first'),
    หมวดหมู่_TO=('ป้ายกำกับ', 'first'),
    หน่วย_TO=('หน่วย', 'first'),
    โอนสินค้า=('โอนสินค้า', 'sum'),
    ผู้ส่ง=('ผู้ส่ง', 'first'),
    สาขา_ผู้รับ_TO=('ผู้รับ', 'first'),
    สถานะ_TO=('สถานะ TO', 'first'),
)

# --- Aggregate GR: primary by TO+product, secondary by GR+product ---
# Collect ALL GR numbers (comma-separated) to handle 2-round receipts
def join_unique(series):
    vals = [str(v) for v in series.dropna().unique() if str(v).startswith('GR')]
    return ', '.join(vals) if vals else None

gr_agg_by_to = gr.groupby(['TO', 'รหัสวัตถุดิบ'], as_index=False).agg(
    วันที่_GR=('วันที่', 'last'),
    GR=('GR', join_unique),
    เติมสินค้า=('เติมสินค้า', 'sum'),
    สาขา_ผู้รับ_GR=('ผู้รับ', 'first'),
    รอบการรับ=('GR', 'nunique'),
)
gr_agg_by_gr = gr.groupby(['GR', 'รหัสวัตถุดิบ'], as_index=False).agg(
    วันที่_GR_alt=('วันที่', 'last'),
    เติมสินค้า_alt=('เติมสินค้า', 'sum'),
)
gr_docs_in_file = set(gr['GR'].dropna().unique())

# =======================================================================
# PATH A: TR-sourced rows — keep EACH TR row individually (no groupby!)
#   This ensures duplicate TRs (same branch+product+date pointing to same
#   TO) each appear as their own output row.
# =======================================================================
tr_clean = tr[tr['TR'].notna() & tr['TR'].str.startswith('TR', na=False)].copy()
tr_clean['ขอโอนสินค้า'] = pd.to_numeric(tr_clean.get('ขอโอนสินค้า', 0), errors='coerce').fillna(0)
# Rename ALL TR columns that would collide with TO columns BEFORE the merge
tr_clean = tr_clean.rename(columns={
    'วันที่ออก TR': 'วันที่_TR',
    'สถานะ TR':     'สถานะWสถานะ_TR',
    'ผู้รับ'       'สาขา_ผู้รับ_TR',
    'ผู้ส่ง':       'ผู้ส่ง_TR',   # avoid ผู้ส่ง_x /ผู้ส่ง_y collision
    'ชื่อ':         'ชื่อ_TR',
    'ป้ายกำกับ':    'ป้ายกำกับ_TR',
    'หน่วย':        'หน่วย_TR',
})

# Join TO data onto each TR row
path_a = tr_clean.merge(
    to_agg[['TO','รหัสวัตถุดิบ',
            'วันที่_TO','TR_from_TO','GR_from_TO',
            'ชื่อ_TO','หมวดหมู่_TO','หน่วย_TO',
            'โอนสินค้า','ผู้ส่ง','สาขา_ผู้รับ_TO','สถานะ_TO']],
    on=['TO','รหัสวัตถุดิบ'], how='left'
)

# Join GR (primary: TO+product)
path_a = path_a.merge(
    gr_agg_by_to[['TO','รหัสวัตถุดิบ','วันที่_GR','GR','เติมสินค้า','รอบการรับ']],
    on=['TO','รหัสวัตถุดิบ'], how='left'
)
path_a['GR'] = path_a['GR'].fillna(path_a.get('GR_from_TO'))

# =======================================================================
# PATH B: TO-only rows — direct transfers with no TR in TR file
# =======================================================================
tr_tos = set(tr_clean['TO'].dropna().astype(str))
path_b = to_agg[~to_agg['TO'].astype(str).isin(tr_tos)].copy()
path_b['วันที่_TR']       = None
path_b['TR']               = None
path_b['สาขา_ผู้รับ_TR']  = None
path_b['สถานะ_TR']        = None
path_b['ขอโอนสินค้า']     = 0.0
path_b['ชื่อ_TR']         = None
path_b['ป้ายกำกับ_TR']    = None
path_b['หน่วย_TR']        = None
path_b['ผู้ส่ง_TR']       = None

path_b = path_b.merge(
    gr_agg_by_to[['TO','รหัสวัตถุดิบ','วันที่_GR','GR','เติมสินค้า','รอบการรับ']],
    on=['TO','รหัสวัตถุดิบ'], how='left'
)
path_b['GR'] = path_b['GR'].fillna(path_b['GR_from_TO'])

# =======================================================================
# UNION both paths into merged
# =======================================================================
UNION_COLS = [
    'วันที่_TR','TR','ขอโอนสินค้า','สถานะ_TR','สาขา_ผู้รับ_TR','ผู้ส่ง_TR',
    'ชื่อ_TR','ป้ายกำกับ_TR','หน่วย_TR',
    'TO','รหัสวัตถุดิบ','GR_from_TO',
    'วันที่_TO','ชื่อ_TO','หมวดหมู่_TO','หน่วย_TO',
    'โอนสินค้า','ผู้ส่ง','สาขา_ผู้รับ_TO','สถานะ_TO',
    'วันที่_GR','GR','เติมสินค้า','รอบการรับ',
]
merged = pd.concat([
    path_a.reindex(columns=UNION_COLS),
    path_b.reindex(columns=UNION_COLS),
], ignore_index=True)

# Fill GR from TO file reference (belt-and-suspenders)
merged['GR'] = merged['GR'].fillna(merged['GR_from_TO'])

# Secondary GR join for rows with no match: GR.GR + product = TO.GR_ref + product
need_secondary = merged['เติมสินค้า'].isna() & merged['GR_from_TO'].notna()
if need_secondary.any():
    sec = merged[need_secondary][['GR_from_TO','รหัสวัตถุดิบ']].copy()
    sec = sec.merge(gr_agg_by_gr, left_on=['GR_from_TO','รหัสวัตถุดิบ'], right_on=['GR','รหัสวัตถุดิบ'], how='left')
    merged.loc[need_secondary, 'เติมสินค้า'] = sec['เติมสินค้า_alt'].values
    merged.loc[need_secondary, 'วันที่_GR'] = sec['วันที่_GR_alt'].values
    merged.loc[need_secondary, 'gr_via_secondary'] = True

# Tag rows where GR reference exists but no product match found
merged['gr_ref_exists_no_match'] = (
    merged['เติมสินค้า'].isna() &
    merged['GR_from_TO'].notna() &
    merged['GR_from_TO'].isin(gr_docs_in_file)
)
merged['gr_ref_not_in_file'] = (
    merged['เติมสินค้า'].isna() &
    merged['GR_from_TO'].notna() &
    ~merged['GR_from_TO'].isin(gr_docs_in_file)
)

# Product name + unit: prefer TO data, fallback to TR data
merged['ชื่อสินค้า'] = merged['ชื่อ_TO'].fillna(merged['ชื่อ_TR'])
merged['หน่วย']     = merged['หน่วย_TO'].fillna(merged['หน่วย_TR'])
merged['ป้ายกำกับ'] = merged['หมวดหมู่_TO'].fillna(merged['ป้ายกำกับ_TR'])

# Receiver branch: prefer TO's ผู้รับ, fallback to TR's ผู้รับ
merged['สาขาผู้รับ'] = merged['สาขา_ผู้รับ_TO'].fillna(merged['สาขา_ผู้รับ_TR'])
# Sender: prefer TO's ผู้ส่ง (authoritative), fallback to TR's ผู้ส่ง
merged['ผู้ส่ง'] = merged['ผู้ส่ง'].fillna(merged['ผู้ส่ง_TR'])

# Fill NaN quantities with 0 for comparison
merged['ขอโอนสินค้า'] = merged['ขอโอนสินค้า'].fillna(0)
merged['โอนสินค้า'] = merged['โอนสินค้า'].fillna(0)
merged['เติมสินค้า'] = merged['เติมสินค้า'].fillna(0)

# --- Analysis note ---
def get_note(row):
    tr_qty = row['ขอโอนสินค้า']
    to_qty = row['โอนสินค้า']
    gr_qty = row['เติมสินค้า']
    status_to = str(row['สถานะ_TO']) if pd.notna(row['สถานะ_TO']) else ''
    status_tr = str(row['สถานะ_TR']) if pd.notna(row['สถานะ_TR']) else ''
    has_tr = pd.notna(row['TR']) and str(row['TR']).startswith('TR')
    has_to = to_qty > 0 or (status_to not in ('', 'nan'))

    # ── TR-level cancellation / rejection (no TO created yet) ──────
    if has_tr and not has_to:
        if 'Canceled' in status_tr or 'Cancelled' in status_tr:
            if 'dest_br' in status_tr:
                return '❌ ยกเลิก TR โดยสาขาปลายทาง (ไม่มี TO)'
            return '❌ ยกเลิก TR (ไม่มี TO)'
        if 'Rejected' in status_tr:
            if 'source_br' in status_tr:
                return '❌ TR ถูกปฏิเสธโดยต้นทาง (ไม่มี TO)'
            return '❌ TR ถูกปฏิเสธ (ไม่มี TO)'
        # TR is valid but no TO was ever created → คลังไม่ได้ส่ง
        return '⚠️ คลังลืมส่ง / ของหมด (มี TR แต่ไม่มี TO)'

    # ── TO-level cancellation ───────────────────────────────────────
    if 'Canceled' in status_to or 'Cancelled' in status_to:
        if 'dest_br' in status_to:
            return '❌ ยกเลิกโดยสาขาปลายทาง'
        return '❌ ยกเลิก'

    # ── TR rejected (TO exists but TR was rejected) ─────────────────
    if 'Rejected' in status_tr:
        return '❌ TR ถูกปฏิเสธ'

    # ── Still in progress: only flag when GR has NOT arrived yet ────
    # (POS sometimes keeps status='Ordered' even after GR is received)
    if status_to == 'Ordered' and gr_qty == 0:
        if has_tr:
            return '⏳ อยู่ระหว่างดำเนินการ (ยังไม่ส่ง)'
        else:
            return '⏳ รอดำเนินการ (ไม่มีใบร้องขอ)'

    # No TR = direct send
    if not has_tr:
        if gr_qty == 0:
            gr_ref = row.get('GR', '')
            gr_ref_exists_no_match = row.get('gr_ref_exists_no_match', False)
            gr_ref_not_in_file = row.get('gr_ref_not_in_file', False)
            if gr_ref_exists_no_match:
                return f'⚠️ ส่งตรง (ไม่มี TR) - มี GR ({gr_ref}) แต่ไม่พบรายการสินค้านี้ใน GR'
            elif gr_ref_not_in_file:
                return f'📋 ส่งตรง (ไม่มี TR) - GR ออกแล้ว ({gr_ref}) แต่ยังไม่ปรากฏในไฟล์รับโอน'
            return '⚠️ ส่งตรง (ไม่มี TR) - ยังไม่รับเข้าระบบ'
        if to_qty == gr_qty:
            return '✅ ส่งตรง (ไม่มี TR) - รับครบ'
        elif gr_qty < to_qty:
            diff = to_qty - gr_qty
            return f'⚠️ ส่งตรง (ไม่มี TR) - รับไม่ครบ (ขาด {diff:.4g} {row["หน่วย"]})'
        return '✅ ส่งตรง (ไม่มี TR)'

    # Has TR - full flow
    if to_qty == 0 and gr_qty == 0:
        return '⏳ สั่งแล้ว ยังไม่ส่ง/รับ'

    notes = []

    # GR ref issues (when gr_qty still 0 after all join attempts)
    if gr_qty == 0 and to_qty > 0:
        gr_ref = row.get('GR', '')
        gr_ref_exists_no_match = row.get('gr_ref_exists_no_match', False)
        gr_ref_not_in_file = row.get('gr_ref_not_in_file', False)
        if gr_ref_exists_no_match:
            notes.append(f'⚠️ มี GR ({gr_ref}) แต่ไม่พบรายการสินค้านี้ใน GR - ตรวจสอบด้วยตนเอง')
        elif gr_ref_not_in_file:
            notes.append(f'📋 GR ออกแล้ว ({gr_ref}) แต่ยังไม่ปรากฏในไฟล์รับโอน')
        else:
            notes.append('📋 ยังไม่รับเข้าระบบ')

    # Check TO vs TR
    if tr_qty > 0 and to_qty < tr_qty:
        diff = tr_qty - to_qty
        notes.append(f'ส่งไม่ครบ (สั่ง {tr_qty:.4g} ส่ง {to_qty:.4g} ขาด {diff:.4g})')
    elif tr_qty > 0 and to_qty > tr_qty:
        diff = to_qty - tr_qty
        notes.append(f'ส่งเกิน (สั่ง {tr_qty:.4g} ส่ง {to_qty:.4g} เกิน {diff:.4g})')

    # Check GR vs TO (skip if GR-issue already noted above)
    gr_issue_already_noted = any('GR' in n and ('ออกแล้ว' in n or 'ไม่พบ' in n) for n in notes)
    if to_qty > 0 and gr_qty == 0 and not gr_issue_already_noted:
        notes.append('📋 ยังไม่รับเข้าระบบ')
    elif gr_qty < to_qty:
        diff = to_qty - gr_qty
        notes.append(f'ของหาย/รับไม่ครบ (ส่ง {to_qty:.4g} รับ {gr_qty:.4g} ขาด {diff:.4g})')
    elif gr_qty > to_qty:
        diff = gr_qty - to_qty
        notes.append(f'รับเกินใบส่ง (ส่ง {to_qty:.4g} รับ {gr_qty:.4g} เกิน {diff:.4g})')

    if not notes:
        return '✅ ปกติ (สั่ง=ส่ง=รับ)'

    joined = ' | '.join(notes)
    # If already has emoji prefix from GR-issue notes, return as-is
    if joined.startswith('⚠️') or joined.startswith('📋') or joined.startswith('✅'):
        return joined
    prefix = '⚠️ ' if 'ขาด' in joined or 'หาย' in joined else '📋 '
    return prefix + joined

merged['หมายเหตุ'] = merged.apply(get_note, axis=1)

# --- Duplicate order detection ---
# Flag: same branch + same product + same TR date ordered more than once
tr_for_dup = tr[tr['TR'].notna() & tr['TR'].str.startswith('TR', na=False)].copy()
tr_for_dup['วันที่ออก TR'] = tr_for_dup['วันที่ออก TR'].astype(str)

# Collect all TR doc numbers per group (branch + product + date)
dup_tr_list = tr_for_dup.groupby(['ผู้รับ', 'รหัสวัตถุดิบ', 'วันที่ออก TR'])['TR'].apply(
    lambda x: sorted(x.dropna().unique().tolist())
).reset_index()
dup_tr_list.columns = ['สาขาผู้รับ_dup', 'รหัสวัตถุดิบ', 'วันที่_TR_dup', 'TR_list']
dup_tr_list = dup_tr_list[dup_tr_list['TR_list'].apply(len) > 1]

merged['วันที่_TR_str'] = merged['วันที่_TR'].astype(str)
# Use TO branch, fallback to TR branch (for TR rows where TO data is missing)
merged['สาขา_dup'] = merged['สาขาผู้รับ'].astype(str)
dup_map = {}  # key -> list of TR numbers
for _, row in dup_tr_list.iterrows():
    key = (row['สาขาผู้รับ_dup'], row['รหัสวัตถุดิบ'], row['วันที่_TR_dup'])
    dup_map[key] = row['TR_list']

def flag_dup(row):
    key = (str(row['สาขา_dup']), str(row['รหัสวัตถุดิบ']), str(row['วันที่_TR_str']))
    if key in dup_map:
        tr_nums = dup_map[key]
        tr_str = ', '.join(tr_nums)
        return f'⚠️ พบสั่งซ้ำ {len(tr_nums)} ใบ: {tr_str}'
    return ''

merged['แจ้งเตือนสั่งซ้ำ'] = merged.apply(flag_dup, axis=1)

# --- Build final output columns ---
if 'รอบการรับ' not in merged.columns:
    merged['รอบการรับ'] = 0
merged['รอบการรับ'] = pd.to_numeric(merged['รอบการรับ'], errors='coerce').fillna(0).astype(int)

out = merged[[
    'วันที่_TR', 'วันที่_TO', 'วันที่_GR',
    'สาขาผู้รับ', 'ผู้ส่ง',
    'รหัสวัตถุดิบ', 'ชื่อสินค้า', 'ป้ายกำกับ', 'หน่วย',
    'TR', 'TO', 'GR',
    'สถานะ_TR', 'สถานะ_TO',
    'ขอโอนสินค้า', 'โอนสินค้า', 'เติมสินค้า',
    'รอบการรับ', 'หมายเหตุ', 'แจ้งเตือนสั่งซ้ำ'
]].copy()

out.columns = [
    'วันที่สั่ง (TR)', 'วันที่ส่ง (TO)', 'วันที่รับ (GR)',
    'สาขาผู้รับ', 'ผู้ส่ง',
    'รหัสวัตถุดิบ', 'ชื่อสินค้า', 'หมวดหมู่', 'หน่วย',
    'เลข TR', 'เลข TO', 'เลข GR',
    'สถานะ TR', 'สถานะ TO',
    'จำนวนสั่ง', 'จำนวนส่ง', 'จำนวนรับ',
    'รอบการรับ GR', 'หมายเหตุ / วิเคราะห์', 'แจ้งเตือนสั่งซ้ำ'
]

# Sort by branch, send date, product
out = out.sort_values(['สาขาผู้รับ', 'วันที่ส่ง (TO)', 'ชื่อสินค้า']).reset_index(drop=True)

print(f'Total rows: {len(out)}')
print(out['หมายเหตุ / วิเคราะห์'].value_counts().head(20))
print(out.head(3).to_string())

# --- Write to Excel with openpyxl ---
wb = openpyxl.Workbook()

# ==================== SHEET 1: รายละเอียด ====================
ws1 = wb.active
ws1.title = 'รายละเอียดทั้งหมด'

HEADER_FILL = PatternFill('solid', start_color='1F4E79')
HEADER_FONT = Font(name='Arial', bold=True, color='FFFFFF', size=10)
NORMAL_FONT = Font(name='Arial', size=9)
BOLD_FONT = Font(name='Arial', bold=True, size=9)

# Color fills for status
FILL_OK = PatternFill('solid', start_color='E2EFDA')       # light green
FILL_WARN = PatternFill('solid', start_color='FFEB9C')     # yellow
FILL_ERR = PatternFill('solid', start_color='FFC7CE')      # red
FILL_CANCEL = PatternFill('solid', start_color='D9D9D9')   # grey
FILL_PENDING = PatternFill('solid', start_color='DDEBF7')  # light blue
FILL_DUP = PatternFill('solid', start_color='FFD966')      # orange-yellow for duplicates
FILL_ALT = PatternFill('solid', start_color='F5F5F5')      # light grey alternating

thin = Side(style='thin', color='BFBFBF')
thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)

headers = list(out.columns)
ws1.append(headers)

# Header style
for col_idx, h in enumerate(headers, 1):
    cell = ws1.cell(row=1, column=col_idx)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thin_border

ws1.row_dimensions[1].height = 30

# Data rows
for r_idx, row in enumerate(out.itertuples(index=False), 2):
    note = str(row[headers.index('หมายเหตุ / วิเคราะห์')])
    dup_warn = str(row[headers.index('แจ้งเตือนสั่งซ้ำ')])
    if dup_warn and dup_warn != 'nan':
        row_fill = FILL_DUP
    elif '✅' in note:
        row_fill = FILL_OK
    elif '❌' in note:
        row_fill = FILL_CANCEL
    elif '⏳' in note:
        row_fill = FILL_PENDING
    elif '⚠️' in note:
        row_fill = FILL_WARN
    else:
        row_fill = None

    for c_idx, val in enumerate(row, 1):
        cell = ws1.cell(row=r_idx, column=c_idx)
        if pd.isna(val) or val == 'nan':
            cell.value = ''
        elif isinstance(val, float) and val == int(val):
            cell.value = int(val)
        else:
            cell.value = val
        cell.font = NORMAL_FONT
        cell.border = thin_border
        cell.alignment = Alignment(vertical='center', wrap_text=False)
        if row_fill:
            cell.fill = row_fill

    ws1.row_dimensions[r_idx].height = 15

# Qty columns: center
qty_cols = [headers.index('จำนวนสั่ง')+1, headers.index('จำนวนส่ง')+1, headers.index('จำนวนรับ')+1]
for r_idx in range(2, len(out)+2):
    for c in qty_cols:
        ws1.cell(row=r_idx, column=c).alignment = Alignment(horizontal='center', vertical='center')

# Column widths
col_widths = {
    'วันที่สั่ง (TR)': 13, 'วันที่ส่ง (TO)': 13, 'วันที่รับ (GR)': 13,
    'สาขาผู้รับ': 20, 'ผู้ส่ง': 16,
    'รหัสวัตถุดิบ': 20, 'ชื่อสินค้า': 28, 'หมวดหมู่': 14, 'หน่วย': 8,
    'เลข TR': 16, 'เลข TO': 16, 'เลข GR': 30,
    'สถานะ TR': 20, 'สถานะ TO': 24,
    'จำนวนสั่ง': 10, 'จำนวนส่ง': 10, 'จำนวนรับ': 10,
    'รอบการรับ GR': 12, 'หมายเหตุ / วิเคราะห์': 50, 'แจ้งเตือนสั่งซ้ำ': 35,
}
for i, col in enumerate(headers, 1):
    ws1.column_dimensions[get_column_letter(i)].width = col_widths.get(col, 14)

ws1.freeze_panes = 'A2'

# Add autofilter
ws1.auto_filter.ref = f'A1:{get_column_letter(len(headers))}{len(out)+1}'

# ==================== SHEET 2: สรุปรายสาขา-วัน ====================
ws2 = wb.create_sheet('สรุปรายสาขา-วัน')

# Build summary per branch + date (using TO date as reference date)
summary = out.copy()
summary['วันที่ส่ง (TO)'] = summary['วันที่ส่ง (TO)'].fillna(summary['วันที่สั่ง (TR)'])

# Separate direct send vs TR-based
def cat(note):
    if '✅' in note: return 'ปกติ'
    if '❌' in note: return 'ยกเลิก'
    if '⏳' in note: return 'รอดำเนินการ'
    return 'มีความผิดปกติ'

summary['สถานะรวม'] = summary['หมายเหตุ / วิเคราะห์'].apply(cat)

# Per branch + day aggregation
grp = summary.groupby(['สาขาผู้รับ', 'วันที่ส่ง (TO)'], as_index=False).agg(
    **{
        'รายการทั้งหมด': ('ชื่อสินค้า', 'count'),
        'จำนวนสั่งรวม': ('จำนวนสั่ง', 'sum'),
        'จำนวนส่งรวม': ('จำนวนส่ง', 'sum'),
        'จำนวนรับรวม': ('จำนวนรับ', 'sum'),
        'รายการปกติ': ('สถานะรวม', lambda x: (x == 'ปกติ').sum()),
        'รายการผิดปกติ': ('สถานะรวม', lambda x: (x == 'มีความผิดปกติ').sum()),
        'รายการยกเลิก': ('สถานะรวม', lambda x: (x == 'ยกเลิก').sum()),
        'รายการรอดำเนินการ': ('สถานะรวม', lambda x: (x == 'รอดำเนินการ').sum()),
    }
)

grp['ส่ง-สั่ง'] = grp['จำนวนส่งรวม'] - grp['จำนวนสั่งรวม']
grp['รับ-ส่ง'] = grp['จำนวนรับรวม'] - grp['จำนวนส่งรวม']

def sum_note(row):
    notes = []
    if row['รายการผิดปกติ'] > 0:
        notes.append(f"⚠️ มีรายการผิดปกติ {int(row['รายการผิดปกติ'])} รายการ")
    if row['รายการยกเลิก'] > 0:
        notes.append(f"❌ ยกเลิก {int(row['รายการยกเลิก'])} รายการ")
    if row['รายการรอดำเนินการ'] > 0:
        notes.append(f"⏳ รอดำเนินการ {int(row['รายการรอดำเนินการ'])} รายการ")
    if not notes:
        return '✅ ทุกรายการปกติ'
    return ' | '.join(notes)

grp['สรุปสถานะ'] = grp.apply(sum_note, axis=1)

# Build detail strings for abnormal and pending items per branch+day
def build_detail(group_df, status_filter):
    rows = group_df[group_df['สถานะรวม'] == status_filter]
    lines = []
    for _, r in rows.iterrows():
        qty_info = f"สั่ง {r['จำนวนสั่ง']:.4g} | ส่ง {r['จำนวนส่ง']:.4g} | รับ {r['จำนวนรับ']:.4g}"
        note_short = r['หมายเหตุ / วิเคราะห์'].replace('⚠️ ','').replace('📋 ','').replace('✅ ','').replace('⏳ ','').replace('❌ ','')
        lines.append(f"• {r['ชื่อสินค้า']} ({qty_info}) → {note_short}")
    return '\n'.join(lines)

detail_abnormal = summary.groupby(['สาขาผู้รับ','วันที่ส่ง (TO)']).apply(
    lambda g: build_detail(g, 'มีความผิดปกติ')
).reset_index(name='รายละเอียดผิดปกติ')

detail_pending = summary.groupby(['สาขาผู้รับ','วันที่ส่ง (TO)']).apply(
    lambda g: build_detail(g, 'รอดำเนินการ')
).reset_index(name='รายละเอียดรอดำเนินการ')

grp = grp.merge(detail_abnormal, on=['สาขาผู้รับ','วันที่ส่ง (TO)'], how='left')
grp = grp.merge(detail_pending, on=['สาขาผู้รับ','วันที่ส่ง (TO)'], how='left')
grp = grp.sort_values(['สาขาผู้รับ', 'วันที่ส่ง (TO)']).reset_index(drop=True)

sum_headers = ['สาขาผู้รับ', 'วันที่ส่ง (TO)', 'รายการทั้งหมด',
               'จำนวนสั่งรวม', 'จำนวนส่งรวม', 'จำนวนรับรวม',
               'ส่ง-สั่ง', 'รับ-ส่ง',
               'รายการปกติ', 'รายการผิดปกติ', 'รายการยกเลิก', 'รายการรอดำเนินการ',
               'สรุปสถานะ', 'รายละเอียดผิดปกติ', 'รายละเอียดรอดำเนินการ']

ws2.append(sum_headers)
for col_idx, h in enumerate(sum_headers, 1):
    cell = ws2.cell(row=1, column=col_idx)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thin_border
ws2.row_dimensions[1].height = 30

detail_cols = {'รายละเอียดผิดปกติ', 'รายละเอียดรอดำเนินการ'}
for r_idx, row_data in enumerate(grp[sum_headers].itertuples(index=False), 2):
    status_note = str(row_data[sum_headers.index('สรุปสถานะ')])
    if '✅' in status_note:
        row_fill = FILL_OK
    elif '⚠️' in status_note:
        row_fill = FILL_WARN
    else:
        row_fill = None

    max_lines = 1
    for c_idx, val in enumerate(row_data, 1):
        col_name = sum_headers[c_idx - 1]
        cell = ws2.cell(row=r_idx, column=c_idx)
        if pd.isna(val) or val == 'nan' or val == '':
            cell.value = ''
        elif isinstance(val, float) and val == int(val):
            cell.value = int(val)
        else:
            cell.value = val
        is_detail = col_name in detail_cols
        cell.font = NORMAL_FONT
        cell.border = thin_border
        cell.alignment = Alignment(vertical='top', wrap_text=is_detail, horizontal='left' if is_detail else 'general')
        if row_fill:
            cell.fill = row_fill
        if is_detail and isinstance(val, str) and val:
            max_lines = max(max_lines, val.count('\n') + 1)

    ws2.row_dimensions[r_idx].height = max(15, max_lines * 14)

sum_col_widths = {
    'สาขาผู้รับ': 22, 'วันที่ส่ง (TO)': 14,
    'รายการทั้งหมด': 13, 'จำนวนสั่งรวม': 13, 'จำนวนส่งรวม': 13, 'จำนวนรับรวม': 13,
    'ส่ง-สั่ง': 10, 'รับ-ส่ง': 10,
    'รายการปกติ': 12, 'รายการผิดปกติ': 14, 'รายการยกเลิก': 13, 'รายการรอดำเนินการ': 16,
    'สรุปสถานะ': 40, 'รายละเอียดผิดปกติ': 70, 'รายละเอียดรอดำเนินการ': 60,
}
for i, col in enumerate(sum_headers, 1):
    ws2.column_dimensions[get_column_letter(i)].width = sum_col_widths.get(col, 14)
ws2.freeze_panes = 'A2'
ws2.auto_filter.ref = f'A1:{get_column_letter(len(sum_headers))}{len(grp)+1}'

# ==================== Save ====================
out_path = 'mnt/ตรวจสอบสินค้าจัดส่งสาขา/ตารางเปรียบเทียบการสั่ง-ส่ง-รับ.xlsx'
wb.save(out_path)
print(f'\nSaved to {out_path}')
print(f'Sheet 1: {len(out)} rows | Sheet 2: {len(grp)} rows')
